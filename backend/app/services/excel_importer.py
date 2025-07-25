import pandas as pd
import io
from typing import List, Optional
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from .csv_importer import create_table_from_dataframe, generate_create_table_sql
from app.services.type_detection import detect_column_type
from app.services.column_utils import build_column_preview_info, generate_table_name_from_filename
from app.services.dataframe_converter import convert_dataframe_types_from_detection


def get_excel_sheets(file_content: bytes) -> List[str]:
    """
    Get list of sheet names from an Excel file
    """
    try:
        workbook = load_workbook(io.BytesIO(file_content), read_only=True)
        return workbook.sheetnames
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")


async def import_excel_to_table(
    db: Session,
    file: UploadFile,
    table_name: str,
    sheet_name: Optional[str] = None,
    create_table: bool = True,
    detect_types: bool = True
) -> dict:
    """
    Import Excel file to database table with type detection
    """
    contents = await file.read()
    
    # Get available sheets
    sheets = get_excel_sheets(contents)
    
    if not sheets:
        raise HTTPException(status_code=400, detail="No sheets found in Excel file")
    
    # Select sheet to import
    if sheet_name and sheet_name not in sheets:
        raise HTTPException(
            status_code=400, 
            detail=f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(sheets)}"
        )
    
    selected_sheet = sheet_name or sheets[0]
    
    # Read the Excel file
    try:
        df = pd.read_excel(io.BytesIO(contents), sheet_name=selected_sheet, engine='openpyxl')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    
    if df.empty:
        raise HTTPException(status_code=400, detail=f"Sheet '{selected_sheet}' is empty")
    
    if not table_name:
        filename = file.filename or "uploaded_table"
        # Generate table name using shared utility
        if len(sheets) > 1:
            table_name = generate_table_name_from_filename(filename, selected_sheet)
        else:
            table_name = generate_table_name_from_filename(filename)
    
    # Detect column types
    column_types = {}
    type_conversions = {}
    
    if detect_types:
        for col in df.columns:
            sql_type, pandas_dtype = detect_column_type(df[col])
            column_types[col] = sql_type
            type_conversions[col] = pandas_dtype
        
        # Convert DataFrame types using shared utility
        df = convert_dataframe_types_from_detection(df, column_types)
    
    # Rename columns to be SQL-safe
    original_columns = list(df.columns)
    df.columns = [str(col).strip().replace(' ', '_').replace('-', '_').lower() for col in df.columns]
    
    if create_table and detect_types:
        # Create mapping with original column names for create_table_from_dataframe
        original_column_types = {original_columns[i]: column_types.get(original_columns[i], "TEXT") 
                                for i in range(len(original_columns))}
        
        # Create table with proper types
        create_table_from_dataframe(db, df, table_name, original_column_types)
        # Insert data
        df.to_sql(table_name, con=db.get_bind(), if_exists='append', index=False, method='multi')
    else:
        # Use pandas default behavior
        df.to_sql(table_name, con=db.get_bind(), if_exists='replace', index=False)
    
    row_count = len(df)
    column_count = len(df.columns)
    
    # Build column info with types
    column_info = []
    for orig_col, new_col in zip(original_columns, df.columns):
        column_info.append({
            "name": new_col,
            "original_name": str(orig_col),
            "type": column_types.get(orig_col, "TEXT") if detect_types else "TEXT"
        })
    
    return {
        "message": f"Successfully imported {row_count} rows from sheet '{selected_sheet}' into table '{table_name}'",
        "table_name": table_name,
        "sheet_name": selected_sheet,
        "available_sheets": sheets,
        "row_count": row_count,
        "column_count": column_count,
        "columns": list(df.columns),
        "column_types": column_info if detect_types else None
    }


async def import_excel_all_sheets(
    db: Session,
    file: UploadFile,
    table_prefix: str = "",
    create_table: bool = True,
    detect_types: bool = True
) -> dict:
    """
    Import all sheets from an Excel file into separate tables
    """
    contents = await file.read()
    sheets = get_excel_sheets(contents)
    
    if not sheets:
        raise HTTPException(status_code=400, detail="No sheets found in Excel file")
    
    if not table_prefix:
        filename = file.filename or "uploaded"
        table_prefix = generate_table_name_from_filename(filename)
    
    results = []
    total_rows = 0
    
    for sheet in sheets:
        try:
            # Create a new file-like object for each sheet import
            file_copy = UploadFile(
                filename=file.filename,
                file=io.BytesIO(contents)
            )
            
            # Generate table name for this sheet
            sheet_table_name = f"{table_prefix}_{sheet.lower().replace(' ', '_')}"
            
            # Import the sheet
            result = await import_excel_to_table(
                db=db,
                file=file_copy,
                table_name=sheet_table_name,
                sheet_name=sheet,
                create_table=create_table,
                detect_types=detect_types
            )
            
            results.append(result)
            total_rows += result["row_count"]
            
        except Exception as e:
            results.append({
                "sheet_name": sheet,
                "error": str(e),
                "status": "failed"
            })
    
    return {
        "message": f"Imported {len([r for r in results if 'error' not in r])} sheets with {total_rows} total rows",
        "sheets_imported": results,
        "total_sheets": len(sheets),
        "total_rows": total_rows
    }


def preview_excel_data(file_content: bytes, sheet_name: Optional[str] = None, rows: int = 10) -> dict:
    """
    Preview Excel file data without importing - matches CSV preview format
    """
    sheets = get_excel_sheets(file_content)
    
    if not sheets:
        raise HTTPException(status_code=400, detail="No sheets found in Excel file")
    
    selected_sheet = sheet_name or sheets[0]
    
    if selected_sheet not in sheets:
        raise HTTPException(
            status_code=400,
            detail=f"Sheet '{selected_sheet}' not found. Available sheets: {', '.join(sheets)}"
        )
    
    # Read the full file to get accurate row count, but limit preview
    df_full = pd.read_excel(io.BytesIO(file_content), sheet_name=selected_sheet, engine='openpyxl')
    df = df_full.head(rows)
    
    # Build columns info in CSV preview format
    columns = []
    column_types_dict = {}
    
    for col in df_full.columns:
        sql_type, _ = detect_column_type(df_full[col])
        column_types_dict[col] = sql_type
        
        # Build column preview info using shared utility
        series = df_full[col]
        column_info = build_column_preview_info(series, str(col), sql_type)
        columns.append(column_info)
    
    # Generate CREATE TABLE SQL
    suggested_table_name = f"sheet_{selected_sheet.lower().replace(' ', '_')}"
    create_table_sql, _, _ = generate_create_table_sql(df_full, suggested_table_name, column_types_dict)
    
    # Get sample data
    sample_data = df.to_dict('records')
    
    return {
        "columns": columns,
        "sample_data": sample_data,
        "total_rows": len(df_full),
        "create_table_sql": create_table_sql,
        "suggested_table_name": suggested_table_name,
        "sheet_name": selected_sheet,
        "available_sheets": sheets
    }